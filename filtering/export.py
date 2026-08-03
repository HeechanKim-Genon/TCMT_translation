# -*- coding: utf-8 -*-
"""산출 — 최종 사전 CSV 3종 + 감사용 엑셀

CSV 는 **원본 엑셀과 같은 12컬럼 스펙**으로 낸다. 파생 컬럼(구조유형·판정
라벨 등)은 감사용 엑셀에만 남긴다.

  data/dictionary.csv             77,497행 / 고유 영문명 55,988
  data/dictionary_sample2000.csv   2,806행 / 고유 영문명  2,000
  data/dictionary_sample2000_one.csv  2,000행 (영문명당 1행)
"""
import os
import sys

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import filter_common as fc

SEED = 20260803   # 표본 재현용


def read(name):
    p = fc.w(name)
    return pd.read_csv(p, dtype=str, keep_default_na=False) if os.path.exists(p) else None


def main():
    allrows = read("stage0_all.csv")
    judged2 = read("stage2_judged.csv")
    final = read("stage5.csv")
    if final is None:
        sys.exit("stage5.csv 가 없다. stage1~5 를 먼저 돌려라.")

    # ── CSV 3종 (12컬럼 스펙)
    out = final[fc.DICT_COLS].sort_values(["영문명", "용어코드"])
    out.to_csv(fc.o("dictionary.csv"), index=False, encoding="utf-8-sig")

    terms = out["영문명"].drop_duplicates()
    pick = set(terms.sample(min(2000, len(terms)), random_state=SEED))
    samp = out[out["영문명"].isin(pick)]
    samp.to_csv(fc.o("dictionary_sample2000.csv"), index=False, encoding="utf-8-sig")
    samp.drop_duplicates("영문명").to_csv(
        fc.o("dictionary_sample2000_one.csv"), index=False, encoding="utf-8-sig")

    print(f"dictionary.csv            {len(out):,}행 / 고유 영문명 {out['영문명'].nunique():,}")
    print(f"dictionary_sample2000.csv {len(samp):,}행 / 고유 영문명 {samp['영문명'].nunique():,}")

    # ── 퍼널
    stages = [("원본", allrows, "보건의료용어표준 V7.0 전체"),
              ("1차", read("stage1.csv"), "A/B/C/G 유형 제거 → 1~3단어만 채택"),
              ("2차", read("stage2.csv"), "합성어인 경우 (구성단어 조합으로 재현되는 것)"),
              ("3차-1", read("stage3.csv"), "KCD 코드로 필터링 (V01~Y98)"),
              ("3차-2", read("stage3b.csv"), "비의학용어를 LLM으로 필터링"),
              ("4차", read("stage4.csv"), "한글명이 긴 설명문·EDI 수가명인 경우"),
              ("5차", final, "사전 없이도 맞게 번역되는 일상어 제거")]
    rows, prev = [], None
    n0 = len(allrows) if allrows is not None else len(final)
    for name, df, desc in stages:
        if df is None:
            continue
        n = len(df)
        rows.append([name, desc, n, "" if prev is None else f"-{prev-n:,}",
                     f"{n/n0*100:.2f}%"])
        prev = n
    funnel = pd.DataFrame(rows, columns=["단계", "기준", "잔존 행수", "증감", "원본 대비"])
    print()
    print(funnel.to_string(index=False))

    # ── 1:N 묶음
    grp = (final.groupby("영문명")
           .agg(한글후보수=("한글명", "nunique"),
                한글후보=("한글명", lambda s: " | ".join(dict.fromkeys(s))))
           .reset_index()
           .sort_values(["한글후보수", "영문명"], ascending=[False, True]))

    # ── 감사용 엑셀
    xlsx = fc.o("filtering_audit.xlsx")
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        funnel.to_excel(xw, sheet_name="00_퍼널", index=False)
        if allrows is not None:
            (allrows["구조유형"].value_counts().rename_axis("구조유형")
             .reset_index(name="행수")
             .to_excel(xw, sheet_name="01_구조유형_분포", index=False))
        if judged2 is not None:
            (judged2.groupby(["구조유형", "2차판정"]).size().reset_index(name="행수")
             .to_excel(xw, sheet_name="02_2차판정_분포", index=False))
        final.to_excel(xw, sheet_name="10_최종채택", index=False)
        (grp[grp["한글후보수"] >= 2]
         .to_excel(xw, sheet_name="11_최종채택_1대N", index=False))
        for sheet, name, cols in [
            ("20_제거_2차합성어", "stage2_judged.csv", ["영문명", "한글명", "2차근거"]),
            ("21_제거_3차KCD", "removed_stage3_kcd.csv", ["영문명", "한글명", "KCD"]),
            ("22_제거_3차비의학", "removed_stage3b_medical.csv", ["영문명", "한글명", "구조유형"]),
            ("23_제거_4차한글명", "removed_stage4_korean.csv", ["영문명", "한글명", "EDI"]),
            ("24_제거_5차일상어", "removed_stage5_trivial.csv", ["영문명", "한글명", "구조유형"]),
        ]:
            df = read(name)
            if df is None:
                continue
            if sheet.startswith("20"):
                df = df[df["2차판정"] == "제거:합성어"]
            df[cols].head(50000).to_excel(xw, sheet_name=sheet, index=False)

    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    for ws in wb.worksheets:
        for col in ws.iter_cols(min_row=1, max_row=1):
            c = col[0]
            best = max((len(str(ws.cell(r, c.column).value or ""))
                        for r in range(1, min(ws.max_row, 200) + 1)), default=10)
            ws.column_dimensions[c.column_letter].width = min(max(best + 2, 10), 60)
        ws.freeze_panes = "A2"
    wb.save(xlsx)
    print(f"\n엑셀: {xlsx}")


if __name__ == "__main__":
    main()
